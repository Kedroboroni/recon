import os
import requests
import json
import numpy as np
import openpyxl
from geopy.distance import geodesic
from PySide6.QtWidgets import  QFileDialog
from PySide6.QtGui import QPixmap, QImage
from PySide6.QtCore import Slot, QUrl, QTimer

#from docx.shared import Pt
#from docx.enum.text import WD_ALIGN_PARAGRAPH
#from docx.shared import RGBColor

#Использую символы: ° ′ ″ для форматирования градусов, минут, секунд соответсвенно

apiKey =  'b25d4bd7-6564-4cb9-8a26-5b1f1c9a414b'
fileName = "testTable.xlsx"

def getCoordinateRequest(namePlace, apiKey):
    try:

        # Вернет кортеж (широта, долгота) в десятичном представлении
        lang = 'ru_RU'
        req = requests.get(f'https://search-maps.yandex.ru/v1/?apikey={apiKey}&text={namePlace}&lang={lang}')
        #print(req.json())
        feature = req.json()['features'][0]# Извлечение координат
        coordinates = feature['geometry']['coordinates']
        #print(f"\n Coordinate: {coordinates[1]}, {coordinates[0]} ")
        return (coordinates[1], coordinates[0])
    
    except:
        return (123, 123)


def transformationCoordinate(latitudeDecimal, longitudeDecimal):
    #вернет словарь {latitudeGeografy:"строкаШироты" longitudeGeografy "строкаДолготы"}
    latitudeGeografy = transformationDecimalOnDegries(latitudeDecimal)
    longitudeGeografy = transformationDecimalOnDegries(longitudeDecimal)

    return {"latitudeGeografy": f"{latitudeGeografy[0]}°{latitudeGeografy[1]}′{latitudeGeografy[2]}″", "longitudeGeografy" : f"{longitudeGeografy[0]}°{longitudeGeografy[1]}′{longitudeGeografy[2]}″"}


def transformationDecimalOnDegries(values):
    abs_decimal = abs(values)
    degrees = int(abs_decimal)
    minutes = int((abs_decimal - degrees) * 60)
    seconds = round((abs_decimal - degrees - minutes / 60) * 3600, 2)

    return (degrees, minutes, seconds)


def filingTableCoordinate(fileName, apiKey, getInfo = None):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    for rowNumber in range(1,sheet.max_row+1):
        namePlace = sheet.cell(row = rowNumber, column = 1).value
        decimalCoordinate = getCoordinateRequest(namePlace, apiKey) #Кортеж с широтой и долготой
        values = (decimalCoordinate[0], decimalCoordinate[1],
                transformationCoordinate(decimalCoordinate[0], decimalCoordinate[1])["latitudeGeografy"],
                transformationCoordinate(decimalCoordinate[0], decimalCoordinate[1])["longitudeGeografy"],
                transformGeografyToGK(decimalCoordinate[0], decimalCoordinate[1])[0],
                transformGeografyToGK(decimalCoordinate[0], decimalCoordinate[1])[1],
                distanceToUkrain([decimalCoordinate[0], decimalCoordinate[1]])) #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

        try:

            geografyCoordinate = transformationCoordinate(decimalCoordinate[0], decimalCoordinate[1]) #Слоаврь со строками
            writeCoordinateToExcel(workbook, fileName, rowNumber,  values)

            if getInfo:
                getInfo.setText("Объекты найдены успешно!")

        except:
            writeCoordinateToExcel(workbook, fileName, rowNumber, 123)


def writeCoordinateToExcel(workbook, fileName, row_number, values):
#def writeCoordinateToExcel(workbook, fileName, row_number, column_number, values value1, value2, value3, value4):
    sheet = workbook.active
    for i, value in enumerate(values):

        sheet.cell(row=row_number, column = i+2).value = value
    #sheet.cell(row=row_number, column=column_number+1).value = value2

    #sheet.cell(row=row_number, column=column_number+2).value = value3
    #sheet.cell(row=row_number, column=column_number+3).value = value4
    workbook.save(fileName)


@Slot()
def openFile(parent):
    options = QFileDialog.Options()
    fileName, _ = QFileDialog.getOpenFileName(parent, options=options)
    if fileName:
        parent.setText(fileName)
        print(parent.text())


@Slot()
def openFolder(parent):
    options = QFileDialog.Options()
    folderName = QFileDialog.getExistingDirectory(parent, options=options)

    if folderName:
        parent.setText(folderName)
        print(os.listdir(parent.text())) #Список объектов в папке


def recordingGeografyСoordinates(fileName, getInfo = None):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    for rowNumber in range(1,sheet.max_row+1):
        try:

            recordGeographyCoordinate = transformationCoordinate(sheet.cell(row = rowNumber, column = 1).value, sheet.cell(row = rowNumber, column = 2).value)
            sheet.cell(row=rowNumber, column=3).value = recordGeographyCoordinate["latitudeGeografy"]
            sheet.cell(row=rowNumber, column=4).value = recordGeographyCoordinate["longitudeGeografy"]

        except:
            sheet.cell(row=rowNumber, column=3).value = 123
            sheet.cell(row=rowNumber, column=4).value = 123

    if getInfo:
        getInfo.setText("Перобразование успшно произошло")

    workbook.save(fileName)


def mergeExcelFiles(nameFolder, nameOutputFile):
    workbook = openpyxl.Workbook() # Создаем новый Excel файл для объединения таблиц
    sheet = workbook.active
    excel_files = [f for f in os.listdir(nameFolder) if f.endswith('.xlsx')] # Получаем список всех xlsx файлов в директории

    for i, fileName in enumerate(excel_files):  # Объединяем все таблицы
        
        workbookNew = openpyxl.load_workbook(os.path.join(nameFolder, fileName))
        sheetNew = workbookNew.active

        if i == 0: #Проверка для устарнения дублирования заглавных строк
            minRow = 1
        else:
            minRow = 2 

        for row in sheetNew.iter_rows(min_row=minRow, values_only=True):  # Копируем данные в активную рабочую область нового Excel файла
            sheet.append(row)
    
    # Сохраняем результат в новый файл
    workbook.save(nameOutputFile)


def getColumnByNumber(nameFile, columnNumber):
    #Достанет из таблицы nameFile, Колонку номер columnNumber
    values = [] #Список значений ячеек колонки
    workbook = openpyxl.load_workbook(nameFile)
    sheet = workbook.active
    values = []

    for row in sheet.iter_rows(min_row=1, max_col = columnNumber, min_col = columnNumber):

        values.extend([cell.value for cell in row])

    return values

def insertInColumnByNumber(nameFile, values, columnNumber):
    #Вставляет значения в колонку по индексу колонки
    try:

        workbook = openpyxl.load_workbook(nameFile)

    except:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    for i, value in enumerate(values):

        sheet.cell(row = i+1, column = columnNumber, value = value) #Вставляем в колонку номер columnNumber, значения в списке values

    workbook.save(nameFile)


def insertFromToWhere(nameFileFrom, nameFileWhere, numberColumnFrom, numberColumnWhere):

    values = getColumnByNumber(nameFileFrom, numberColumnFrom)
    insertInColumnByNumber(nameFileWhere, values, numberColumnWhere)


def createNewTableByOldDate(nameFile, columns, nameFileOld):
    #Создаст новую таблицу с указанными колонками из старой
    for i, column in enumerate(columns):

        insertFromToWhere(nameFileOld, nameFile, column, i+1)


def generalAnalys(pathFolder, pathFileMerge, pathFileEdit, pathEditNew):

    mergeExcelFiles(pathFolder, pathFileMerge)
    print("Слили таблицы в одну")

    createNewTableByOldDate(pathFileEdit, (6,7,8, 9, 10), pathFileMerge)
    print("Создали новую таблицу с нужными столбцами")

    numberColumn = 3 #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    valueColumn = "Россия" #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    removeRowsForSort(valueColumn, pathFileEdit, numberColumn) # Теперь таблица только с Россией
    listRepitsValues = getColumnByNumber(pathFileEdit, 5) # Список столбца с названием объекта !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    ListCountry = ["Россия" for i in listRepitsValues] # Список Россия
    listRegions = getColumnByNumber(pathFileEdit, 4) # Список80 столбца с Регионом !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    dictCountRepitsValues = countMeet(listRepitsValues) # Слоаврь с повторениями
    
    #insertInColumnByNumber(pathEditNew, listRepitsValues, 1) #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    #insertInColumnByNumber(pathEditNew, ListCountry, 2) #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    #insertInColumnByNumber(pathEditNew, listRegions, 3) #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    listCount = [dictCountRepitsValues[i] for i in dictCountRepitsValues] #Список количесвта встречаемых объектов
    listObj = list(dictCountRepitsValues.keys())
    insertInColumnByNumber(pathEditNew, listObj, 1)
    insertInColumnByNumber(pathEditNew, listCount, 2)
    

def removeRowsForSort(valueColumn, pathFile, numberColumn):
    #Удаляет все не свопадения по значению в колонке, возвращает длинну получившейся таблицы
    i = 1
    workbook = openpyxl.load_workbook(pathFile)
    sheet = workbook.active
    listValuesChoesColumn = getColumnByNumber(pathFile, numberColumn)

    while sheet.cell(row = i, column = 1).value:
        
        try:
            while listValuesChoesColumn[i-1] != valueColumn:
                
                listValuesChoesColumn.remove(listValuesChoesColumn[i-1])
                sheet.delete_rows(i)
                #mainworkbook.save(pathFile)
                print(i)
            
                listValuesChoesColumn[i]
        except:
                listValuesChoesColumn.remove(listValuesChoesColumn[i-1])
                sheet.delete_rows(i)
                print("Выходим из цикла")
                break
        i = i+1
    workbook.save(pathFile)
    return len(listValuesChoesColumn)


def countMeet(listValue):
    #Подсчитывает количесвто повторений значения из спискаЮ возвращает словарь со значением по ключу (количесвто повторений)
    dictRepits = {name: 1 for  name in listValue}

    for i in dictRepits:

        count = 0

        for j in listValue:

            if i == j:
                count = count+1

        dictRepits[i] = count

    return dictRepits


def transformGeografyToGK(lat, lon, zone = None):
    # Номер зоны Гаусса-Крюгера (если точка рассматривается в системе
    # координат соседней зоны, то номер зоны следует присвоить вручную)
    if zone == None:
        zone = int(lon/6.0+1)

    # Импорт математических функций
    from math import sin, cos, tan, pi

    # Параметры эллипсоида Красовского
    a = 6378245.0          # Большая (экваториальная) полуось
    b = 6356863.019        # Малая (полярная) полуось
    e2 = (a**2-b**2)/a**2  # Эксцентриситет
    n = (a-b)/(a+b)        # Приплюснутость

    # Параметры зоны Гаусса-Крюгера
    F = 1.0                   # Масштабный коэффициент
    Lat0 = 0.0                # Начальная параллель (в радианах)
    Lon0 = (zone*6-3)*pi/180  # Центральный меридиан (в радианах)
    N0 = 0.0                  # Условное северное смещение для начальной параллели
    E0 = zone*1e6+500000.0    # Условное восточное смещение для центрального меридиана

    # Перевод широты и долготы в радианы
    Lat = lat*pi/180.0
    Lon = lon*pi/180.0

    # Вычисление переменных для преобразования
    v = a*F*(1-e2*(sin(Lat)**2))**-0.5
    p = a*F*(1-e2)*(1-e2*(sin(Lat)**2))**-1.5
    n2 = v/p-1
    M1 = (1+n+5.0/4.0*n**2+5.0/4.0*n**3)*(Lat-Lat0)
    M2 = (3*n+3*n**2+21.0/8.0*n**3)*sin(Lat-Lat0)*cos(Lat+Lat0)
    M3 = (15.0/8.0*n**2+15.0/8.0*n**3)*sin(2*(Lat-Lat0))*cos(2*(Lat+Lat0))
    M4 = 35.0/24.0*n**3*sin(3*(Lat-Lat0))*cos(3*(Lat+Lat0))
    M = b*F*(M1-M2+M3-M4)
    I = M+N0
    II = v/2*sin(Lat)*cos(Lat)
    III = v/24*sin(Lat)*(cos(Lat))**3*(5-(tan(Lat)**2)+9*n2)
    IIIA = v/720*sin(Lat)*(cos(Lat)**5)*(61-58*(tan(Lat)**2)+(tan(Lat)**4))
    IV = v*cos(Lat)
    V = v/6*(cos(Lat)**3)*(v/p-(tan(Lat)**2))
    VI = v/120*(cos(Lat)**5)*(5-18*(tan(Lat)**2)+(tan(Lat)**4)+14*n2-58*(tan(Lat)**2)*n2)

    # Вычисление северного и восточного смещения (в метрах)
    X = I+II*(Lon-Lon0)**2+III*(Lon-Lon0)**4+IIIA*(Lon-Lon0)**6
    Y = E0+IV*(Lon-Lon0)+V*(Lon-Lon0)**3+VI*(Lon-Lon0)**5
    return (f"{X:.4f}", f"{Y:.4f}")


def recordingGeografyСoordinatesToGK(lon, lat, getInfo, zone = None):
    
    X = transformGeografyToGK(lon, lat, zone)[0]
    Y = transformGeografyToGK(lon, lat, zone)[1]
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    for rowNumber in range(1,sheet.max_row+1):
        try:

            recordGeographyCoordinate = transformGeografyToGK(sheet.cell(row = rowNumber, column = 1).value, sheet.cell(row = rowNumber, column = 2).value, zone)
            sheet.cell(row=rowNumber, column=3).value = recordGeographyCoordinate[0]
            sheet.cell(row=rowNumber, column=4).value = recordGeographyCoordinate[1]

        except:
            sheet.cell(row=rowNumber, column=3).value = 123
            sheet.cell(row=rowNumber, column=4).value = 123

    if getInfo:
        getInfo.setText("Перобразование успшно произошло")

    workbook.save(fileName)

def distanceToUkrain(listCoordinate):
    try:
        file = parseFileToTuples('border.txt')
        arrFile = np.array(file)
        coordinates = arrFile.reshape(-1, 2)
        distances = [calculateDistance(listCoordinate[0], listCoordinate[1], float(coord[0]), float(coord[1]))
                for coord in coordinates]
        return min(distances)
    except:
        return "ОШИБКА"
        


def parseFileToTuples(file_name):
    with open(file_name, 'r') as file:
        content = file.read()
    
    listCoordinate = content.split("\n")
    filingList = [tuple(i.split(",")) for i in listCoordinate if i !=""]
    return filingList


def calculateDistance(lat1, lon1, lat2, lon2):
    # Создаем объект Distance с координатами первой точки
    point1 = (lat1, lon1)
    
    # Создаем объект Distance с координатами второй точки
    point2 = (lat2, lon2)
    
    # Вычисляем расстояние между точками
    distance = geodesic(point1, point2).kilometers
    
    return round(distance, 2)